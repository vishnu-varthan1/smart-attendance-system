#!/usr/bin/env python3
"""
Helper utilities for the attendance system
"""

import os
import csv
import logging
from datetime import datetime, date, timedelta
from werkzeug.utils import secure_filename
import uuid

# Try to import bleach for HTML sanitization
try:
    import bleach
    BLEACH_AVAILABLE = True
except ImportError:
    BLEACH_AVAILABLE = False
    bleach = None

def setup_logging():
    """Setup logging configuration"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('attendance_system.log'),
            logging.StreamHandler()
        ]
    )

def sanitize_input(text, allow_basic_html=False):
    """Sanitize user input to prevent XSS attacks"""
    if not text:
        return text
    
    # Convert to string if not already
    text = str(text).strip()
    
    if BLEACH_AVAILABLE:
        if allow_basic_html:
            # Allow basic formatting tags
            allowed_tags = ['b', 'i', 'u', 'em', 'strong', 'br', 'p']
            allowed_attributes = {}
        else:
            # Strip all HTML tags
            allowed_tags = []
            allowed_attributes = {}
        
        # Clean the text
        cleaned_text = bleach.clean(
            text, 
            tags=allowed_tags, 
            attributes=allowed_attributes,
            strip=True
        )
        
        return cleaned_text
    else:
        # Fallback: basic HTML escaping
        import html
        return html.escape(text)

def validate_leave_request_data(data):
    """Validate and sanitize leave request data"""
    errors = []
    
    # Required fields
    required_fields = ['student_id', 'leave_type', 'start_date', 'end_date', 'reason']
    for field in required_fields:
        if not data.get(field):
            errors.append(f"{field.replace('_', ' ').title()} is required")
    
    # Sanitize text fields
    if data.get('reason'):
        data['reason'] = sanitize_input(data['reason'])
        if len(data['reason']) < 10:
            errors.append("Reason must be at least 10 characters long")
        if len(data['reason']) > 500:
            errors.append("Reason must be less than 500 characters")
    
    if data.get('leave_type'):
        data['leave_type'] = sanitize_input(data['leave_type'])
        # Validate against allowed leave types
        allowed_types = ['Sick', 'Personal', 'Family', 'Academic', 'Other']
        if data['leave_type'] not in allowed_types:
            errors.append("Invalid leave type selected")
    
    # Validate dates
    if data.get('start_date') and data.get('end_date'):
        try:
            from datetime import datetime
            start = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
            end = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
            
            if end < start:
                errors.append("End date cannot be before start date")
            
            # Check if dates are not too far in the past or future
            today = date.today()
            if start < today - timedelta(days=7):
                errors.append("Start date cannot be more than 7 days in the past")
            
            if end > today + timedelta(days=365):
                errors.append("End date cannot be more than 1 year in the future")
                
        except ValueError:
            errors.append("Invalid date format")
    
    return errors, data

def save_uploaded_file(file, upload_folder, prefix=""):
    """Save uploaded file with secure filename"""
    try:
        if not file or not file.filename:
            return None
        
        # Create upload folder if it doesn't exist
        os.makedirs(upload_folder, exist_ok=True)
        
        # Generate secure filename
        filename = secure_filename(file.filename)
        if not filename:
            filename = f"{prefix}{uuid.uuid4().hex[:8]}.jpg"
        else:
            name, ext = os.path.splitext(filename)
            filename = f"{prefix}{name}_{uuid.uuid4().hex[:8]}{ext}"
        
        filepath = os.path.join(upload_folder, filename)
        file.save(filepath)
        
        return filepath
        
    except Exception as e:
        logging.error(f"Error saving file: {str(e)}")
        return None

def export_attendance_to_csv(records):
    """Export attendance records to CSV"""
    try:
        filename = f"attendance_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join('exports', filename)
        
        os.makedirs('exports', exist_ok=True)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            writer.writerow([
                'Date', 'Student ID', 'Student Name', 'Time In', 'Status', 
                'Department', 'Year', 'Section', 'Marked By'
            ])
            
            # Write records
            for record in records:
                writer.writerow([
                    record.date.strftime('%Y-%m-%d') if record.date else '',
                    record.student.student_id if record.student else '',
                    record.student.name if record.student else '',
                    record.time_in.strftime('%H:%M:%S') if record.time_in else '',
                    record.status,
                    record.student.department if record.student else '',
                    record.student.year if record.student else '',
                    record.student.section if record.student else '',
                    getattr(record, 'marked_by', 'System')
                ])
        
        return filepath
        
    except Exception as e:
        logging.error(f"Error exporting to CSV: {str(e)}")
        return None

def export_attendance_to_excel(records):
    """Export attendance records to Excel with formatting"""
    try:
        # Try to import openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logging.warning("openpyxl not available, falling back to CSV export")
            return export_attendance_to_csv(records)
        
        filename = f"attendance_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join('exports', filename)
        
        os.makedirs('exports', exist_ok=True)
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Records"
        
        # Define headers
        headers = [
            'Date', 'Student ID', 'Student Name', 'Time In', 'Status', 
            'Department', 'Year', 'Section', 'Confidence', 'Marked By'
        ]
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write and style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data rows
        for row_idx, record in enumerate(records, 2):
            data = [
                record.date.strftime('%Y-%m-%d') if record.date else '',
                record.student.student_id if record.student else '',
                record.student.name if record.student else '',
                record.time_in.strftime('%H:%M:%S') if record.time_in else '',
                record.status,
                record.student.department if record.student else '',
                record.student.year if record.student else '',
                record.student.section if record.student else '',
                f"{record.confidence_score:.2f}" if record.confidence_score else '',
                getattr(record, 'marked_by', 'System')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Color code status
                if col == 5:  # Status column
                    if value == 'Present':
                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    elif value == 'Absent':
                        cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    elif value == 'Late':
                        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary at the bottom
        summary_row = len(records) + 3
        ws.cell(row=summary_row, column=1, value="Summary:").font = Font(bold=True)
        
        total_records = len(records)
        present_count = len([r for r in records if r.status == 'Present'])
        absent_count = len([r for r in records if r.status == 'Absent'])
        late_count = len([r for r in records if r.status == 'Late'])
        
        ws.cell(row=summary_row + 1, column=1, value=f"Total Records: {total_records}")
        ws.cell(row=summary_row + 2, column=1, value=f"Present: {present_count}")
        ws.cell(row=summary_row + 3, column=1, value=f"Absent: {absent_count}")
        ws.cell(row=summary_row + 4, column=1, value=f"Late: {late_count}")
        
        # Save workbook
        wb.save(filepath)
        
        return filepath
        
    except Exception as e:
        logging.error(f"Error exporting to Excel: {str(e)}")
        # Fallback to CSV if Excel export fails
        return export_attendance_to_csv(records)

def generate_attendance_summary(records):
    """Generate attendance summary statistics"""
    try:
        total_records = len(records)
        
        if total_records == 0:
            return {
                'total_records': 0,
                'present_count': 0,
                'absent_count': 0,
                'late_count': 0,
                'present_percentage': 0,
                'absent_percentage': 0,
                'late_percentage': 0
            }
        
        present_count = len([r for r in records if r.status == 'Present'])
        absent_count = len([r for r in records if r.status == 'Absent'])
        late_count = len([r for r in records if r.status == 'Late'])
        
        return {
            'total_records': total_records,
            'present_count': present_count,
            'absent_count': absent_count,
            'late_count': late_count,
            'present_percentage': round((present_count / total_records) * 100, 1),
            'absent_percentage': round((absent_count / total_records) * 100, 1),
            'late_percentage': round((late_count / total_records) * 100, 1)
        }
        
    except Exception as e:
        logging.error(f"Error generating summary: {str(e)}")
        return {}

def validate_student_data(data):
    """Validate and sanitize student registration data"""
    errors = []
    
    required_fields = ['student_id', 'name', 'email', 'department', 'year', 'section']
    
    for field in required_fields:
        if not data.get(field):
            errors.append(f"{field.replace('_', ' ').title()} is required")
    
    # Sanitize text fields
    text_fields = ['student_id', 'name', 'email', 'phone', 'department', 'year', 'section']
    for field in text_fields:
        if data.get(field):
            data[field] = sanitize_input(data[field])
    
    # Validate email format (basic)
    email = data.get('email', '')
    if email and '@' not in email:
        errors.append("Invalid email format")
    
    # Validate student ID format (basic)
    student_id = data.get('student_id', '')
    if student_id and len(student_id) < 3:
        errors.append("Student ID must be at least 3 characters")
    
    # Validate name
    name = data.get('name', '')
    if name and len(name) < 2:
        errors.append("Name must be at least 2 characters")
    
    return errors

def create_directory_structure():
    """Create necessary directory structure"""
    directories = [
        'static/uploads',
        'student_images',
        'exports',
        'database',
        'logs'
    ]
    
    for directory in directories:
        os.makedirs(directory, exist_ok=True)

def get_attendance_status(student_id, date_obj=None):
    """Get attendance status for a student on a specific date"""
    if date_obj is None:
        date_obj = date.today()
    
    # This would typically query the database
    # For now, return a placeholder
    return 'Unknown'